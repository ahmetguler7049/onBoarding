from django.shortcuts import render, redirect
from django.urls import reverse
from django.contrib.auth import authenticate, login
from .forms import LoginForm, ForgetForm, ResetPasswordForm

from app.models import *

from django.contrib import messages
from django.utils.safestring import mark_safe
from django.utils.http import urlsafe_base64_decode, urlsafe_base64_encode

import json
import urllib.request
import urllib.parse
from django.conf import settings

from django.contrib.auth.decorators import login_required
from .tokens import account_activation_token
from sendgrid.helpers.mail import Mail
from sendgrid import SendGridAPIClient
from django.contrib.sites.shortcuts import get_current_site
from openpyxl import *
from django.utils.encoding import force_bytes, force_text


def recaptcha_kontrol(request):
    recaptcha_response = request.POST.get('g-recaptcha-response')
    url = 'https://www.google.com/recaptcha/api/siteverify'
    values = {
        'secret': settings.GOOGLE_RECAPTCHA_SECRET_KEY,
        'response': recaptcha_response
    }
    data = urllib.parse.urlencode(values).encode()
    req = urllib.request.Request(url, data=data)
    response = urllib.request.urlopen(req)
    result = json.loads(response.read().decode())
    return result


@login_required(login_url="user_login_view")
def index(request):
    return render(request, "index.html")


def user_login_view(request):
    form = LoginForm(request.POST or None)
    context = {
        'form': form,
    }

    if "user_login" in request.POST:

        if form.is_valid():
            result = recaptcha_kontrol(request)

            if result['success']:
                email = form.cleaned_data.get('email')
                password = form.cleaned_data.get('password')

                try:
                    findUser = User._default_manager.get(email__iexact=email)
                    if findUser is not None:
                        if findUser.check_password(password):
                            user_login_check = authenticate(email=email, password=password)
                            login(request, user_login_check)
                            return redirect("home")
                        else:
                            messages.error(request, mark_safe('Geçersiz Şifre!'))
                except User.DoesNotExist:
                    messages.error(request, mark_safe('E-posta adresi bulunamadı!'))
            else:
                messages.error(request, mark_safe('Geçersiz reCAPTCHA. Lütfen tekrar deneyin.'))
        else:
            messages.error(request, mark_safe('Tüm alanları eksiksiz bir şekilde doldurduğunuzdan emin olun!'))

    return render(request, "accounts/login.html", context=context)


# TODO: Admin Login View'a göre formu özelleştir.

def admin_login_view(request):
    form = LoginForm(request.POST or None)
    context = {
        'form': form,
    }

    if request.method == "POST":

        if form.is_valid():
            result = recaptcha_kontrol(request)

            if result['success']:
                email = form.cleaned_data.get('email')
                password = form.cleaned_data.get('password')

                if not form.cleaned_data.get('remember_me'):
                    request.session.set_expiry(0)

                try:
                    findUser = User._default_manager.get(email__iexact=email)
                    if findUser is not None:
                        if findUser.check_password(password):
                            user_login_check = authenticate(username=findUser, password=password)
                            login(request, user_login_check)
                            return redirect("home")
                        else:
                            messages.error(request, mark_safe('Geçersiz Şifre!'))
                except User.DoesNotExist:
                    messages.error(request, mark_safe('E-posta adresi bulunamadı!'))
            else:
                messages.error(request, mark_safe('Geçersiz reCAPTCHA. Lütfen tekrar deneyin.'))
        else:
            messages.error(request, mark_safe('Tüm alanları eksiksiz bir şekilde doldurduğunuzdan emin olun!'))

    return render(request, "accounts/admin_login.html", context=context)


def forget_password_view(request):
    form = ForgetForm(request.POST or None)

    if request.method == "POST":

        if form.is_valid():
            email = form.cleaned_data.get("email")
            user = User.objects.get(email__iexact=email)

            domain = get_current_site(request).domain
            email_body = {
                'domain': domain,
                'uid': urlsafe_base64_encode(force_bytes(user.pk)),
                'token': account_activation_token.make_token(user),
            }

            link = reverse('reset_password', kwargs={
                'uidb64': email_body['uid'], 'token': email_body['token']})
            activate_url = 'http://' + domain + link

            message = Mail(
                from_email='accelerator@university4society.com',
                to_emails=email,
                subject='LETS Academy Yeni Parolanı Belirle!',
                html_content='Merhaba,\n'
                             'LETS Academy üzerinden şifre değişikliği talebin alındı. {} adresine gidip yeni şifreni '
                             'oluşturabilirsin.'.format(activate_url)
            )

            try:
                sg = SendGridAPIClient('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')
                response = sg.send(message)
                # print(response.status_code)
                # print(response.body)
                # print(response.headers)
            except Exception as e:
                print(e.message)

            messages.success(request, mark_safe("Şifre yenileme e-postasını gönderdik. E-postanı kontrol etmeyi unutma!"))
            return redirect("forget_password")
    return render(request, "accounts/forget_password.html", {"form": form})


def reset_password_view(request, uidb64, token):

    form = ResetPasswordForm(request.POST or None)
    user_id = force_text(urlsafe_base64_decode(uidb64))
    user = User.objects.get(pk=user_id)
    msg = None

    if request.method == "POST":
        if form.is_valid():
            password = form.cleaned_data.get('password')
            password_check = form.cleaned_data.get('password_check')

            if password != password_check:
                msg = "Şifreler eşleşmiyor!"
                messages.error(request, mark_safe(msg))

            elif password == password_check and len(password) < 6:
                msg = "Şifreniz çok kısa!"
                messages.error(request, mark_safe(msg))

            elif account_activation_token.check_token(user, token):
                user.set_password(raw_password=password)
                user.save()
                msg = 'Şifreniz başarıyla değiştirildi, yeni şifrenizi kullanabilirsiniz.'
                messages.success(request, mark_safe(msg))
                return redirect('login')
    return render(request, "accounts/reset-password.html", {"form": form, "msg": msg})


# def register_user(request):
#     msg = None
#     if request.method == "POST":
#         form = SignUpForm(request.POST)
#         if form.is_valid():
#             form.save()
#             msg = 'User created - please <a href="/login">login</a>.'
#             # return redirect("/login/")
#
#         else:
#             msg = 'Form is not valid'
#     else:
#         form = SignUpForm()
#
#     return render(request, "accounts/register.html", {"form": form, "msg": msg})


@login_required
def Bulk_Add_Users(request):
    if 'bulk_users' in request.POST:
        excel_file = request.FILES.get('excel_file', False)
        user = User.objects.get(id=request.user.id)
        firm = Firm.objects.get(firm_name__iexact=user.firm.firm_name)

        firm.bulk_file = excel_file
        firm.save()
        book = load_workbook(str(settings.MEDIA_ROOT) + '/excel/{}'.format(excel_file))
        sheet = book.active

        # Sending mails for per user one by one.
        email_list = []
        for row in range(1, sheet.max_row + 1):
            to_email = str(sheet.cell(row=row, column=1).value)
            email_list.append(to_email)
        for to_email in email_list:
            if not User.objects.filter(email=to_email).exists():
                newuser = User.objects.create_user(email=to_email, firm=firm)
                newuser.is_active = False
                newuser.save()

                domain = get_current_site(request).domain
                email_body = {
                    'domain': domain,
                    'uid': urlsafe_base64_encode(force_bytes(newuser.pk)),
                    'token': account_activation_token.make_token(newuser),
                }

                link = reverse('complete_profile_view', kwargs={
                    'uidb64': email_body['uid'], 'token': email_body['token']})
                activate_url = 'http://' + domain + link

                message = Mail(
                    from_email='accelerator@university4society.com',
                    to_emails=to_email,
                    subject='LETS Academy Hesap Aktivasyonu',
                    html_content='Merhaba,\n'
                                 'Hesabını aktive etmek için {} adresine gidip profilini tamamlayabilirsin.'.format(activate_url)
                )
                try:
                    sg = SendGridAPIClient('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')
                    response = sg.send(message)
                    # print(response.status_code)
                    # print(response.body)
                    # print(response.headers)
                except Exception as e:
                    print(e.message)
            else:
                pass

        return redirect('Bulk_Add_Users')

    return render(request, 'accounts/bulk_add.html')
